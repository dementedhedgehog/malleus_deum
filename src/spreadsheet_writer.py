import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ability_name_font = Font(size=12, italic=False, bold=True)
archetype_name_font = Font(size=12, italic=False, bold=True)

# create alignment style
wrap_alignment = Alignment(wrap_text=True)
                           
blue_fill = PatternFill(start_color='BED3EB', fill_type='solid')



def write_summary_to_spreadsheet(spreadsheet_fname, ability_groups, archetypes):

    # 
    workbook = Workbook()
    
    for sheet_name in workbook.get_sheet_names():
        workbook.remove_sheet(workbook.get_sheet_by_name(sheet_name))

    abilities_sheet = workbook.create_sheet(index = 0, title = "Abilities")
    #archetypes_sheet = workbook.create_sheet(index = 0, title = "Archetypes")

    # grab the active worksheet
    #worksheet = workbook.active

    # Data can be assigned directly to cells
    #worksheet['A1'] = 42

    # Rows can also be appended
    #worksheet.append([1, 2, 3])

    # Python types will automatically be converted
    #xsworksheet['A2'] = datetime.datetime.now()

    write_ability_groups_to_sheet(abilities_sheet, ability_groups, archetypes)
    
    # Save the file
    workbook.save(spreadsheet_fname)
    return



def write_ability_groups_to_sheet(sheet, ability_groups, archetypes):

    # layout constants
    LEFT_MARGIN = 2 # columns
    LEVEL_MARGIN = LEFT_MARGIN + 1 # columns

    TOP_MARGIN = 1 # rows
    #TOP_LABEL_ROW = TOP_MARGIN + 1
    #LABELS_ROW = TOP_LABEL_ROW + 1
    LABELS_ROW = TOP_MARGIN + 1
    ABILITIES_ROW = LABELS_ROW
    ABILITY_GROUP_VSPACE = 2 # rows
    LINE_VSPACE = 1 # rows

    # empty leftmost col
    #sheet.column_dimensions['A'].width = NARROW

    rd = sheet.row_dimensions[LABELS_ROW] # 3]
    #rd = sheet.row_dimensions[3]
    rd.height = 45
    
    # labels    
    row = LABELS_ROW
    col = LEFT_MARGIN
    cell = sheet.cell(row = row, column = col)
    cell.value = "Abilities"

    col += 1
    cell = sheet.cell(row = row, column = col)
    cell.value = "Level"

    col += 1
    cell = sheet.cell(row = row, column = col)
    cell.value = "G/M/L/P/T"

    col += 1
    cell = sheet.cell(row = row, column = col)
    cell.value = "Prereqs"

    # row = TOP_LABEL_ROW
    col += 2    
    # cell = sheet.cell(row = row, column = col)
    # cell.value = "Archetypes"

    # put the archetype names in a row at the top.
    row = ABILITIES_ROW
    for archetype in archetypes:
        cell = sheet.cell(row = row, column = col)
        cell.value = str(archetype.get_title())
        cell.font = archetype_name_font
        cell.alignment = wrap_alignment
        col += 1

    # Fill in the ability information
    row = ABILITIES_ROW
    for ability_group in ability_groups:
        col = LEFT_MARGIN
        cell = sheet.cell(row = row, column = col)
        cell.value = str(ability_group.get_title())
        cell.font = ability_name_font

        for ability in ability_group:
            col = LEFT_MARGIN
            row += LINE_VSPACE
            
            # ability title
            cell = sheet.cell(row = row, column = col)
            cell.value = str(ability.get_title())

            for ability_level in ability.get_levels():
                col = LEVEL_MARGIN

                # ability level
                cell = sheet.cell(row = row, column = col)
                cell.value = str(ability_level.get_level_number())

                # 
                ability_cost = ("%s/%s/%s/%s (%s/%s/%s)" % (
                    ability_level.get_default_general(),
                    ability_level.get_default_martial(),
                    ability_level.get_default_lore(),
                    ability_level.get_default_magical(),
                    ability_level.get_successes(),
                    ability_level.get_attempts(),
                    ability_level.get_failures()))
                col += 1
                cell = sheet.cell(row = row, column = col)
                cell.value = ability_cost

                # prerequisites 
                col += 1
                cell = sheet.cell(row = row, column = col)
                prereqs = ", ".join([str(p) for p in ability_level.get_prerequisites()])
                cell.value = ",".join([str(prereq) for prereq 
                                       in ability_level.get_prerequisites()])

                # put the archetype entries.
                col += 2
                for archetype in archetypes:
                    cell = sheet.cell(row = row, column = col)
                    col += 1

                    if col % 2 == 0:
                        cell.fill = blue_fill
                    
                    modified_ability = archetype.get_modified_ability(ability.get_id())

                    mal = modified_ability.get_modified_ability_level(
                        ability_level.get_level_number())

                    if not mal.is_enabled():
                        cell.value = "n/a"

                    else:
                        lore = mal.get_lore_points()
                        martial = mal.get_martial_points()
                        general = mal.get_general_points()
                        magical = mal.get_magical_points()
                        total = lore + martial + general + magical
                        successes = mal.get_mastery_successes()
                        attempts = mal.get_mastery_attempts()
                        failures = mal.get_mastery_failures()
                        
                        points = "%i/%i/%i/%i (%s/%s/%s) = %i " % (general,
                                                                   martial,
                                                                   lore,
                                                                   magical,
                                                                   successes,
                                                                   attempts,
                                                                   failures,
                                                                   total)
                        cell.value = points
                row += 1

        row += ABILITY_GROUP_VSPACE



    # column widths
    NARROW = 2.0
    STANDARD = 6.0
    WIDE = 22.0
    #sheet.column_dimensions[col].width = STANDARD
    sheet.column_dimensions['A'].width = NARROW
    sheet.column_dimensions['B'].width = WIDE
    sheet.column_dimensions['C'].width = STANDARD
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['F'].width = NARROW
    sheet.column_dimensions['G'].width = WIDE
    sheet.column_dimensions['H'].width = WIDE
    sheet.column_dimensions['I'].width = WIDE
    sheet.column_dimensions['J'].width = WIDE
    sheet.column_dimensions['K'].width = WIDE
    sheet.column_dimensions['L'].width = WIDE
    sheet.column_dimensions['M'].width = WIDE


    
    # sheet.column_dimensions['L'].width = NARROW
    # sheet.column_dimensions['K'].width = 12
    
    # sheet.column_dimensions['M'].width = 15
    # sheet.column_dimensions['N'].width = 15
    # sheet.column_dimensions['O'].width = 15
    # sheet.column_dimensions['P'].width = 15
    # sheet.column_dimensions['Q'].width = 15

	

    # column_widths = []
    # for row in sheet.iter_rows():
    #     print row
    #     for i, cell in enumerate(row):
    #         if len(column_widths) > i:
    #             if len(cell) > column_widths[i]:
    #                 column_widths[i] = len(cell)
    #             else:
    #                 column_widths += [len(cell)]

    # for i, column_width in enumerate(column_widths):
    #     worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
    return

